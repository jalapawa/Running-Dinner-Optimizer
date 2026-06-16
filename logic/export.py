from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from datetime import datetime
from urllib.parse import quote
from api.geocode import geocode
import colorsys

import matplotlib.pyplot as plt

def export(manager):

    distribution = manager.get_optimum()
    groups = manager.get_groups()
    mapping = manager.get_map()
    afterpartyloc = "Trichtergasse 24"
    afterpartycoords = geocode(afterpartyloc, "Aachen")

    groups = sorted(groups, key=lambda x: x.id)

    wb = Workbook()
    ws = wb.active

    for col in "ABCDEFGHIJKLMNOPQRSTUVWXYZ":
        ws.column_dimensions[col].width = 30
    
    yellow_fill = PatternFill(
        fill_type="solid",
        fgColor="FFFF00"
    )

    purple_fill = PatternFill("solid", fgColor="A02B93")
    pink_fill = PatternFill("solid", fgColor="F1CEEE")

    # Header row
    ws.append([
        "Teamname",
        "Adresse der Küche",
        "Wie erreicht man euch im Notfall (Telefonnummer)",
        "Hinweise an die zu bekochenden Gäste (z.B. wie findet man eure Küche, Klingelschild, Stockwerk, Zimmernummer etc.)",
        "Hinweis an die Köche (Allergien, Unverträglichkeiten etc.)",
        "Link zur Route"
    ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = yellow_fill


    # Group data
    for g in groups:
        ws.append([
            g.teamname,
            g.address,
            g.phone,
            g.guestInfo,
            g.cookInfo,
            "Route"
        ])

    for i in range(len(groups)):
        addresses = [
            groups[i].address,
            groups[groups[i].route[0]-1].address,
            groups[groups[i].route[1]-1].address,
            groups[groups[i].route[2]-1].address,
            afterpartyloc
        ]

        link = "https://www.google.com/maps/dir/" + "/".join(
            quote(addr) for addr in addresses if addr is not None
        )

        ws[f"F{i+2}"].hyperlink = link

    # Two empty rows
    ws.append([])
    ws.append([])

    team1 = ["Vorspeise bei:"]
    team2 = [""]
    team3 = [""]
    # Distribution data
    for key, value in list(distribution.items())[:(len(distribution) // 3)]:
        team1.extend([mapping[key]])
        team2.extend([mapping[value[0]]])
        team3.extend([mapping[value[1]]])
    ws.append(team1)
    for cell in ws[ws.max_row]:
        cell.fill = purple_fill
    ws.append(team2)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill
    ws.append(team3)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill

    ws.append([])

    team1 = ["Hauptspeise bei:"]
    team2 = [""]
    team3 = [""]
    # Distribution data
    for key, value in list(distribution.items())[(len(distribution) // 3) : (2 * len(distribution) // 3)]:
        team1.extend([mapping[key]])
        team2.extend([mapping[value[0]]])
        team3.extend([mapping[value[1]]])
    ws.append(team1)
    for cell in ws[ws.max_row]:
        cell.fill = purple_fill
    ws.append(team2)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill
    ws.append(team3)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill

    ws.append([])

    team1 = ["Nachspeise bei:"]
    team2 = [""]
    team3 = [""]
    # Distribution data
    for key, value in list(distribution.items())[(2 * len(distribution) // 3):]:
        team1.extend([mapping[key]])
        team2.extend([mapping[value[0]]])
        team3.extend([mapping[value[1]]])
    ws.append(team1)
    for cell in ws[ws.max_row]:
        cell.fill = purple_fill
    ws.append(team2)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill
    ws.append(team3)
    for cell in ws[ws.max_row]:
        cell.fill = pink_fill

    for cell in ws["A"]:
        cell.font = cell.font.copy(bold=True)

    date = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")

    wb.save(f"results/Rudischwimmt Einteilung-{date}.xlsx")


    
    # routes = [[groups[i].coords,
    #             groups[groups[i].route[0]-1].coords,
    #             groups[groups[i].route[1]-1].coords,
    #             groups[groups[i].route[2]-1].coords,
    #             #afterpartycoords
    #         ] for i in range(len(groups))]
    

    # colors = [colorsys.hsv_to_rgb(i / len(groups), 1.0, 1.0) for i in range(len(groups))]

    # lines = []

    # for i, (route, color) in enumerate(zip(routes, colors)):
    #     lats = [p[0] for p in route]
    #     lons = [p[1] for p in route]

    #     line, = plt.plot(
    #         lons,
    #         lats,
    #         color=color,
    #         label=f"Vehicle {i+1}",
    #         linewidth=2
    #     )
    #     lines.append(line)

    # leg = plt.legend()

    # for legline, origline in zip(leg.get_lines(), lines):
    #     legline.set_picker(5)

    #     def on_pick(event):
    #         line = lines[leg.get_lines().index(event.artist)]
    #         line.set_visible(not line.get_visible())
    #         plt.draw()

    #     plt.gcf().canvas.mpl_connect("pick_event", on_pick)

    # plt.show()